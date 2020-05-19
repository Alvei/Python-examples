""" cell_update.py
    Corrects costs in produce sale spreadsheet. """

import openpyxl

inputfile = "ProduceSales.xlsx"
outputfile = "updatedProduceSales.xlsx"

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 30.07,
                 'Celery': 10.19,
                 'Lemon': 10.27}

print(f"\nReading input file: {inputfile} ...")
wb = openpyxl.load_workbook(inputfile)
sheet = wb["Sheet"]

for row_num in range(2, sheet.max_row): # Skip the 1st row
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

print(f"Writing to output file: {outputfile} ...")
wb.save(outputfile)
