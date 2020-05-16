""" basic_examples.py """
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def print_sheet_area(area):
    """ Print each row at a time. """
    for rows_of_cells in area:
        for cell in rows_of_cells:
            print(f"{cell.coordinate}\t{cell.value}")
        print("---END OF ROW ---")


wb = openpyxl.load_workbook('example.xlsx')
print(f"\nWork book:]\t{type(wb)}")

print(f"Sheet names:\t{wb.sheetnames}")
sheet = wb["Sheet3"]
print(f"Picked sheet:\t{type(sheet)} with name: {sheet.title}")

# Normally sheet at index 0 if you create an empy one
anothersheet = wb.active
print(f"Active sheet:\t{anothersheet.title}")

# Getting cells
################
sheet = wb["Sheet1"]
# By name
print(f"sheet[A1]: {sheet['A1'].value}")
c = sheet['B1']
print(f"{c.coordinate} is Row {c.row}, Column {c.column} has the value {c.value}")
b = sheet.cell(row=1, column=2)
print(f"{b.coordinate} is Row {b.row}, Column {b.column} has the value {b.value}")

# Knowing there are 8 rows, printing every other row
for index in range(1, 8, 2):
    print(f"{index} {sheet.cell(row=index, column=2).value}")

# To find-out the size of the sheet
max_row = sheet.max_row
max_col = sheet.max_column
print(f'Size of sheet: {max_row} by {max_col}')

# Converting column letters and numbers
print(f'Size of sheet: {max_row} by {get_column_letter(max_col)}')
col = "AA"
print(f"Get column index for {col}: {column_index_from_string('AA')}\n")

# Printing a column
for cell in sheet['B']:
    print(cell.value)

# Printing an area from a sheet using slicing, which creates a generator object
print_sheet_area(sheet['A1':'C3'])