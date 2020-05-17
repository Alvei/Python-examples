""" read_census_excel.py
    program that reads census data from excel and does some summation.
    It then saves the resulting data in a python file that can be imported as a module.
    census2010.py keeps the info in a dictionary all_data with key:value as state:county"""
import openpyxl, pprint
from openpyxl.utils import get_column_letter, column_index_from_string

print("Opening Workbook...")
filename = 'censuspopdata.xlsx'
wb = openpyxl.load_workbook(filename)
print(f"Sheet names:\t{wb.sheetnames}")
sheet = wb['Population by Census Tract']

county_data = {}
print("Reading rows...")

for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has dat for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # Make sure the key for this state exists
    county_data.setdefault(state, {})
    # Make sure the key for this county in this state exists
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census track, so increment by one
    county_data[state][county]['tracts'] += 1

    # Increase the country pop by the pop in this census tract
    county_data[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of county_data to i
print("Writing results...")
resultfile =  open('census2010.py', 'w')
resultfile.write('all_data = ' + pprint.pformat(county_data))
resultfile.close()
print("Done.")

